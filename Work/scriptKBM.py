from selenium import webdriver
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC

import win32com.client
import xlwings as xw
import datetime
import openpyxl
import time
import os


def getDataFromExcel(pathToData, dictPage):
    wb = openpyxl.load_workbook(filename = pathToData)
    names = list(dictPage.keys())

    cellsValues = []
    for name in names:
        my_range = wb.defined_names[name]

        # if this contains a range of cells then the destinations attribute is not None
        dests = my_range.destinations # returns a generator of (worksheet title, cell range) tuples

        for title, coord in dests:
            ws = wb[title]
            if isinstance(ws[coord].value, datetime.datetime):
                #print(name, ws[coord].value.strftime('%d.%m.%Y'), type(ws[coord].value.strftime('%d.%m.%Y')))
                cellsValues.append(ws[coord].value.strftime('%d.%m.%Y'))
            elif isinstance(ws[coord].value, type(None)):
                #print(name, ws[coord].value, type(ws[coord].value))
                cellsValues.append('')
            else:
                #print(name, ws[coord].value, type(str(ws[coord].value)))
                cellsValues.append(str(ws[coord].value))
    return dict(zip(names, cellsValues))


def login(driver):
    elem = driver.find_element(By.NAME, "mainForm:j_idt43:0:j_idt132")
    elem.clear()
    elem.send_keys("")

    elem = driver.find_element(By.NAME, "mainForm:j_idt43:1:j_idt135")
    elem.clear()
    elem.send_keys("")
    elem.send_keys(Keys.RETURN)
    assert "Неверный пароль" not in driver.page_source


def fillTextBox(driver, source, data):
    elem = driver.find_element(By.NAME, source)
    elem.clear()
    elem.send_keys(data)


def fillDefaultDriverLicense(driver, source1, source2):
    WebDriverWait(driver, 100).until(EC.element_to_be_clickable((By.ID, source1+'_label'))).click()#ВУ Документ выдан
    time.sleep(0.5)
    WebDriverWait(driver, 100).until(EC.element_to_be_clickable((By.ID, source1+'_1'))).click()#РФ
    time.sleep(1)
    WebDriverWait(driver, 100).until(EC.element_to_be_clickable((By.ID, source2+'_label'))).click()#ВУ Тип документа ВУ
    time.sleep(0.5)
    WebDriverWait(driver, 100).until(EC.element_to_be_clickable((By.ID, source2+'_1'))).click()#ВУ РФ
    time.sleep(1)


def fillIdentityCard(driver, source1, source2, data):
    WebDriverWait(driver, 100).until(EC.element_to_be_clickable((By.ID, source1))).click()#УЛ Документ выдан
    time.sleep(0.5)
    WebDriverWait(driver, 100).until(EC.element_to_be_clickable((By.ID, source1.replace('label', '1')))).click()#РФ
    time.sleep(1)
    WebDriverWait(driver, 100).until(EC.element_to_be_clickable((By.ID, source2))).click()#Тип документа(УЛ)
    time.sleep(0.5)
    WebDriverWait(driver, 100).until(EC.element_to_be_clickable((By.ID, source2.replace('label', str(data))))).click()#Тип документа(УЛ)
    time.sleep(1)


def fillPolicy(driver, source1, data):
    WebDriverWait(driver, 100).until(EC.element_to_be_clickable((By.ID, source1))).click()#Полис договора Серия
    time.sleep(0.5)
    WebDriverWait(driver, 100).until(EC.element_to_be_clickable((By.ID, source1.replace('label', str(data))))).click()#Полис договора Серия
    time.sleep(1)


def fillWithRestrictions(driver, dictPage, dictExcel):
    assert "КБМ" in driver.title
    login(driver)

    time.sleep(3)

    keyList = list(dictPage.keys())

    fillTextBox(driver, dictPage[keyList[0]], dictExcel[keyList[0]])  # Расчетная дата
    fillTextBox(driver, dictPage[keyList[1]], dictExcel[keyList[1]])  # Фамилия
    fillTextBox(driver, dictPage[keyList[2]], dictExcel[keyList[2]])  # Имя
    fillTextBox(driver, dictPage[keyList[3]], dictExcel[keyList[3]])  # Отчество
    fillTextBox(driver, dictPage[keyList[4]], dictExcel[keyList[4]])  # Дата рождения

    fillDefaultDriverLicense(driver, 'mainForm:j_idt2178', 'mainForm:j_idt2286')  # ВУ Документ выдан, Тип документа

    fillTextBox(driver, dictPage[keyList[5]], dictExcel[keyList[5]])  # ВУ Серия
    fillTextBox(driver, dictPage[keyList[6]], dictExcel[keyList[6]])  # ВУ Номер

    fillIdentityCard(driver, 'mainForm:j_idt2661_label', dictPage[keyList[7]],
                     typesDoc[dictExcel[keyList[7]]])  # УЛ Документ выдан, Тип документа

    fillTextBox(driver, dictPage[keyList[8]], dictExcel[keyList[8]])  # УЛ Серия
    fillTextBox(driver, dictPage[keyList[9]], dictExcel[keyList[9]])  # УЛ Номер

    fillPolicy(driver, dictPage[keyList[10]], typesPolicy[dictExcel[keyList[10]]])  # Полис договора Серия

    fillTextBox(driver, dictPage[keyList[11]], dictExcel[keyList[11]])  # Полис договора Номер
    ##################################################################################################################
    if dictExcel[keyList[16]] != '':
        WebDriverWait(driver, 100).until(
            EC.element_to_be_clickable((By.ID, 'mainForm:j_idt3656_toggler'))).click()  # Раскрыть[Доп]
        time.sleep(1.5)

        fillTextBox(driver, dictPage[keyList[17]], dictExcel[keyList[17]])  # [Доп]Фамилия

        fillTextBox(driver, dictPage[keyList[18]], dictExcel[keyList[18]])  # [Доп]Имя

        fillTextBox(driver, dictPage[keyList[19]], dictExcel[keyList[19]])  # [Доп]Отчество

        fillDefaultDriverLicense(driver, 'mainForm:j_idt4178',
                                 'mainForm:j_idt4286')  # [Доп]ВУ Документ выдан, Тип документа

        fillTextBox(driver, dictPage[keyList[20]], dictExcel[keyList[20]])  # [Доп]ВУ Серия
        fillTextBox(driver, dictPage[keyList[21]], dictExcel[keyList[21]])  # [Доп]ВУ Номер

        fillIdentityCard(driver, 'mainForm:j_idt4706_label', dictPage[keyList[22]],
                         typesDoc[dictExcel[keyList[22]]])  # УЛ Документ выдан, Тип документа

        fillTextBox(driver, dictPage[keyList[23]], dictExcel[keyList[23]])  # [Доп]УЛ Серия
        fillTextBox(driver, dictPage[keyList[24]], dictExcel[keyList[24]])  # [Доп]УЛ Номер

        fillTextBox(driver, dictPage[keyList[16]],
                    dictExcel[keyList[16]])  # [Доп]ID !!!Почему-то сбрасывается иногда!!!


def submitDataReturnResult(driver):
    WebDriverWait(driver, 100).until(EC.element_to_be_clickable((By.ID, 'mainForm:j_idt5848'))).click()#Направить запрос
    time.sleep(2)
    elem = driver.find_element(By.CLASS_NAME, 'lastRequestId')
    if elem.text == '':
        elem = driver.find_element(By.ID, 'mainForm:j_idt5867')
        return elem.text
    else:
        return elem.text


dictPage = {
    'kbm_d': 'mainForm:j_idt1524_input',
    'famil': 'mainForm:j_idt1709',
    'name': 'mainForm:j_idt1817',
    'otchestvo': 'mainForm:j_idt1925',
    'born_date': 'mainForm:j_idt2069_input',

    'VU_ser': 'mainForm:j_idt2359',
    'VU_num': 'mainForm:j_idt2517',

    'doc_type': 'mainForm:j_idt2769_label',
    'doc_ser': 'mainForm:j_idt2843',
    'doc_num': 'mainForm:j_idt2946',

    'pol_ser': 'mainForm:j_idt3526_label',
    'pol_num': 'mainForm:j_idt3597',

    'vin': 'mainForm:j_idt3058',
    'num_kuzov': 'mainForm:j_idt3166',
    'num_shassi': 'mainForm:j_idt3274',
    'gos_zn': 'mainForm:j_idt3386',
    #####[Доп]
    'old_id': 'mainForm:j_idt3754',

    'old_famil': 'mainForm:j_idt3818',
    'old_name': 'mainForm:j_idt3926',
    'old_otchestvo': 'mainForm:j_idt4034',

    'old_vu_ser': 'mainForm:j_idt4358',
    'old_vu_num': 'mainForm:j_idt4461',

    'old_doc_type': 'mainForm:j_idt4919_label',
    'old_doc_ser': 'mainForm:j_idt5061',
    'Old_doc_num': 'mainForm:j_idt5259'
}

typesDoc = {
    'Паспорт гражданина РФ': 1,
    'Вид на жительство': 2,
    'Военный билет офицера запаса': 3,
    'Военный билет солдата (матроса, сержанта, старшины)': 4,
    'Временное удостоверение личности гражданина РФ': 5,
    'Дипломатический паспорт гражданина РФ': 6,
    'Другие документы': 7,
    'Загранпаспорт гражданина РФ': 8,
    'Загранпаспорт гражданина СССР': 9,
    'Иностранный паспорт': 10,
    'Иные документы, выдаваемые органами МВД': 11,
    'Паспорт гражданина СССР': 12,
    'Паспорт Минморфлота': 13,
    'Паспорт моряка': 14,
    'Свидетельство о регистрации ходатайства иммигранта о признании его беженцем': 15,
    'Свидетельство о рождении': 16,
    'Свидетельство о рождении, выданное уполномоченным органом иностранного государства': 17,
    'Свидетельство о смерти': 18,
    'Справка об освобождении из места лишения свободы': 19,
    'Удостоверение беженца в РФ': 20,
    'Удостоверение личности офицера': 21,
    '': 0
}
typesPolicy = {
    'ААА': 1,
    'ВВВ': 2,
    'ССС': 3,
    'ЕЕЕ': 4,
    'ХХХ': 5,
    'ККК': 6,
    'МММ': 7,
    'ННН': 8,
    'РРР': 9,
    'ААВ': 10,
    'ААС': 11,
    'ТТТ': 12,
    '': 0
}

os.environ['MOZ_HEADLESS'] = '1'
s = Service(r'C:\Users\Zhmurin.Roman\Anaconda3\Scripts\geckodriver.exe')
driver = webdriver.Firefox(service=s)
driver.get("http://172.19.3.41/") # Открытие сайта
pathToData = os.getcwd() + '\\' + xw.books.active.name

dictExcel = getDataFromExcel(pathToData, dictPage)

fillWithRestrictions(driver, dictPage, dictExcel)
requestId = submitDataReturnResult(driver)
driver.close()

ExcelApp = win32com.client.GetActiveObject("Excel.Application")
ExcelApp.Visible = True

workbook = ExcelApp.ActiveWorkbook
try:
    requestId = int(requestId)
    ExcelApp.Range('AR1048574').Value = requestId
    ExcelApp.Range('CA1048574').Value = ''
    ExcelApp.Range('A1048576').Value = 'Запрос обработан успешно'
except ValueError:
    ExcelApp.Range('AR1048574').Value = ''
    ExcelApp.Range('CA1048574').Value = ''
    ExcelApp.Range('A1048576').Value = requestId

